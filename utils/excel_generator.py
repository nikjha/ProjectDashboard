from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import os


class ExcelGenerator:
    def __init__(self):
        self.workbook = Workbook()
        self.default_sheet = self.workbook.active
        self.default_sheet.title = "Report"
        self.styles = {
            'header': {
                'font': Font(bold=True, color="FFFFFF"),
                'fill': PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid"),
                'alignment': Alignment(horizontal="center", vertical="center")
            },
            'data': {
                'alignment': Alignment(vertical="center")
            },
            'date': {
                'number_format': 'DD-MMM-YYYY'
            }
        }

    def create_sheet(self, title):
        """Create a new worksheet"""
        return self.workbook.create_sheet(title=title[:31])  # Excel sheet name limit

    def add_headers(self, sheet, headers):
        """Add headers to a worksheet"""
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            cell = sheet[f"{col_letter}1"]
            cell.value = header
            for attr, value in self.styles['header'].items():
                setattr(cell, attr, value)

    def add_data(self, sheet, data, start_row=2):
        """Add data rows to a worksheet"""
        for row_num, row_data in enumerate(data, start_row):
            for col_num, cell_value in enumerate(row_data, 1):
                col_letter = get_column_letter(col_num)
                cell = sheet[f"{col_letter}{row_num}"]
                cell.value = cell_value
                for attr, value in self.styles['data'].items():
                    setattr(cell, attr, value)

                # Apply special formatting for dates
                if isinstance(cell_value, (datetime, str)) and "-" in str(cell_value):
                    try:
                        datetime.strptime(str(cell_value), "%Y-%m-%d")
                        for attr, value in self.styles['date'].items():
                            setattr(cell, attr, value)
                    except ValueError:
                        pass

    def auto_size_columns(self, sheet):
        """Auto-size columns to fit content"""
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    value = str(cell.value) if cell.value else ""
                    if len(value) > max_length:
                        max_length = len(value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column_letter].width = adjusted_width

    def add_borders(self, sheet):
        """Add borders to all cells"""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = thin_border

    def save(self, file_path):
        """Save the workbook to a file"""
        # Ensure directory exists
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        self.workbook.save(file_path)

    def generate_report(self, report_data):
        """
        Generate a complete report from structured data
        report_data = {
            'sheets': [
                {
                    'title': 'Tasks',
                    'headers': ['ID', 'Title', 'Status'],
                    'data': [(1, 'Task 1', 'Pending'), ...],
                    'formats': {
                        'columns': {
                            3: {'number_format': '0.00%'}  # Format column 3 as percentage
                        }
                    }
                },
                ...
            ],
            'file_name': 'report.xlsx'
        }
        """
        for sheet_info in report_data['sheets']:
            sheet = self.create_sheet(sheet_info['title'])
            self.add_headers(sheet, sheet_info['headers'])
            self.add_data(sheet, sheet_info['data'])

            # Apply custom column formats if specified
            if 'formats' in sheet_info and 'columns' in sheet_info['formats']:
                for col_num, col_format in sheet_info['formats']['columns'].items():
                    col_letter = get_column_letter(col_num)
                    for cell in sheet[col_letter]:
                        for attr, value in col_format.items():
                            setattr(cell, attr, value)

            self.auto_size_columns(sheet)
            self.add_borders(sheet)

        # Remove default sheet if not used
        if len(self.workbook.worksheets) > 1 and not any(cell.value for cell in self.default_sheet["A1:Z100"]):
            self.workbook.remove(self.default_sheet)

        self.save(report_data['file_name'])
        return report_data['file_name']