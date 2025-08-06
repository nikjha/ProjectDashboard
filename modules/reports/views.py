from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QTableWidget, QTableWidgetItem, QHeaderView, QMessageBox,
    QComboBox, QDateEdit, QGroupBox, QCheckBox
)
from PyQt5.QtCore import Qt, QDate
from PyQt5.QtGui import QIcon
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import os


class ReportsView(QWidget):
    def __init__(self, db_connections, parent=None):
        super().__init__(parent)
        self.db = db_connections

        self.init_ui()

    def init_ui(self):
        self.layout = QVBoxLayout()

        # Filter controls
        filter_group = QGroupBox("Report Filters")
        filter_layout = QVBoxLayout()

        # Date range
        date_layout = QHBoxLayout()
        date_layout.addWidget(QLabel("From:"))
        self.from_date = QDateEdit()
        self.from_date.setDate(QDate.currentDate().addMonths(-1))
        self.from_date.setCalendarPopup(True)
        date_layout.addWidget(self.from_date)

        date_layout.addWidget(QLabel("To:"))
        self.to_date = QDateEdit()
        self.to_date.setDate(QDate.currentDate())
        self.to_date.setCalendarPopup(True)
        date_layout.addWidget(self.to_date)

        filter_layout.addLayout(date_layout)

        # Report type
        self.report_type = QComboBox()
        self.report_type.addItems([
            "Projects Summary",
            "Tasks Report",
            "Bugs Report",
            "Activities Report",
            "Comprehensive Report"
        ])
        filter_layout.addWidget(self.report_type)

        # Options for comprehensive report
        self.include_tasks = QCheckBox("Include Tasks")
        self.include_tasks.setChecked(True)
        self.include_bugs = QCheckBox("Include Bugs")
        self.include_bugs.setChecked(True)
        self.include_activities = QCheckBox("Include Activities")
        self.include_activities.setChecked(True)

        options_layout = QHBoxLayout()
        options_layout.addWidget(self.include_tasks)
        options_layout.addWidget(self.include_bugs)
        options_layout.addWidget(self.include_activities)
        filter_layout.addLayout(options_layout)

        # Generate button
        generate_btn = QPushButton("Generate Report")
        generate_btn.clicked.connect(self.generate_report)
        filter_layout.addWidget(generate_btn)

        filter_group.setLayout(filter_layout)
        self.layout.addWidget(filter_group)

        # Report preview table
        self.report_table = QTableWidget()
        self.report_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.report_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.layout.addWidget(self.report_table)

        self.setLayout(self.layout)

        # Connect signals
        self.report_type.currentIndexChanged.connect(self.update_options_visibility)
        self.update_options_visibility()

    def update_options_visibility(self):
        """Show/hide options based on report type"""
        is_comprehensive = self.report_type.currentText() == "Comprehensive Report"
        self.include_tasks.setVisible(is_comprehensive)
        self.include_bugs.setVisible(is_comprehensive)
        self.include_activities.setVisible(is_comprehensive)

    def generate_report(self):
        """Generate and display the selected report"""
        report_type = self.report_type.currentText()
        from_date = self.from_date.date().toString("yyyy-MM-dd")
        to_date = self.to_date.date().toString("yyyy-MM-dd")

        try:
            if report_type == "Projects Summary":
                self.generate_projects_summary(from_date, to_date)
            elif report_type == "Tasks Report":
                self.generate_tasks_report(from_date, to_date)
            elif report_type == "Bugs Report":
                self.generate_bugs_report(from_date, to_date)
            elif report_type == "Activities Report":
                self.generate_activities_report(from_date, to_date)
            elif report_type == "Comprehensive Report":
                self.generate_comprehensive_report(from_date, to_date)

            QMessageBox.information(self, "Success", "Report generated successfully")
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to generate report: {str(e)}")

    def set_user_role(self, role):
        """Set the user role (currently no role-based restrictions in reports)"""
        self.current_role = role
        # No restrictions applied for reports view

    def generate_projects_summary(self, from_date, to_date):
        """Generate projects summary report"""
        cursor = self.db['sqlite'].cursor()
        cursor.execute('''
                       SELECT p.id,
                              p.name,
                              p.status,
                              COUNT(DISTINCT t.id) as task_count,
                              COUNT(DISTINCT b.id) as bug_count
                       FROM projects p
                                LEFT JOIN tasks t ON p.id = t.project_id
                                LEFT JOIN bugs b ON p.id = b.project_id
                       WHERE p.start_date BETWEEN ? AND ?
                       GROUP BY p.id, p.name, p.status
                       ORDER BY p.name
                       ''', (from_date, to_date))

        projects = cursor.fetchall()

        # Display in table
        self.report_table.setColumnCount(5)
        self.report_table.setHorizontalHeaderLabels([
            "ID", "Project Name", "Status", "Tasks", "Bugs"
        ])

        self.report_table.setRowCount(len(projects))
        for row_idx, project in enumerate(projects):
            for col_idx, value in enumerate(project):
                item = QTableWidgetItem(str(value))
                self.report_table.setItem(row_idx, col_idx, item)

        # Generate Excel
        self.generate_excel_report(
            "Projects Summary",
            ["ID", "Project Name", "Status", "Tasks", "Bugs"],
            projects
        )

    def generate_tasks_report(self, from_date, to_date):
        """Generate tasks report"""
        cursor = self.db['sqlite'].cursor()
        cursor.execute('''
                       SELECT t.id,
                              p.name     as project,
                              t.title,
                              t.status,
                              t.priority,
                              t.due_date,
                              u.username as assigned_to
                       FROM tasks t
                                LEFT JOIN projects p ON t.project_id = p.id
                                LEFT JOIN users u ON t.assigned_to = u.id
                       WHERE t.created_at BETWEEN ? AND ?
                       ORDER BY t.due_date, t.priority DESC
                       ''', (from_date, to_date))

        tasks = cursor.fetchall()

        # Display in table
        self.report_table.setColumnCount(7)
        self.report_table.setHorizontalHeaderLabels([
            "ID", "Project", "Title", "Status", "Priority", "Due Date", "Assigned To"
        ])

        self.report_table.setRowCount(len(tasks))
        for row_idx, task in enumerate(tasks):
            for col_idx, value in enumerate(task):
                item = QTableWidgetItem(str(value) if value else "")
                self.report_table.setItem(row_idx, col_idx, item)

        # Generate Excel
        self.generate_excel_report(
            "Tasks Report",
            ["ID", "Project", "Title", "Status", "Priority", "Due Date", "Assigned To"],
            tasks
        )

    def generate_bugs_report(self, from_date, to_date):
        """Generate bugs report"""
        cursor = self.db['sqlite'].cursor()
        cursor.execute('''
                       SELECT b.id,
                              p.name as project,
                              b.title,
                              b.severity,
                              b.status,
                              b.created_at
                       FROM bugs b
                                LEFT JOIN projects p ON b.project_id = p.id
                       WHERE b.created_at BETWEEN ? AND ?
                       ORDER BY b.severity DESC, b.created_at DESC
                       ''', (from_date, to_date))

        bugs = cursor.fetchall()

        # Display in table
        self.report_table.setColumnCount(6)
        self.report_table.setHorizontalHeaderLabels([
            "ID", "Project", "Title", "Severity", "Status", "Reported On"
        ])

        self.report_table.setRowCount(len(bugs))
        for row_idx, bug in enumerate(bugs):
            for col_idx, value in enumerate(bug):
                item = QTableWidgetItem(str(value) if value else "")
                self.report_table.setItem(row_idx, col_idx, item)

        # Generate Excel
        self.generate_excel_report(
            "Bugs Report",
            ["ID", "Project", "Title", "Severity", "Status", "Reported On"],
            bugs
        )

    def generate_activities_report(self, from_date, to_date):
        """Generate activities report"""
        cursor = self.db['sqlite'].cursor()
        cursor.execute('''
                       SELECT id,
                              title,
                              type,
                              status,
                              created_at
                       FROM extra_activities
                       WHERE created_at BETWEEN ? AND ?
                       ORDER BY created_at DESC
                       ''', (from_date, to_date))

        activities = cursor.fetchall()

        # Display in table
        self.report_table.setColumnCount(5)
        self.report_table.setHorizontalHeaderLabels([
            "ID", "Title", "Type", "Status", "Requested On"
        ])

        self.report_table.setRowCount(len(activities))
        for row_idx, activity in enumerate(activities):
            for col_idx, value in enumerate(activity):
                item = QTableWidgetItem(str(value) if value else "")
                self.report_table.setItem(row_idx, col_idx, item)

        # Generate Excel
        self.generate_excel_report(
            "Activities Report",
            ["ID", "Title", "Type", "Status", "Requested On"],
            activities
        )

    def generate_comprehensive_report(self, from_date, to_date):
        """Generate comprehensive report with multiple sheets"""
        # Create Excel workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Add sheets based on selections
        if self.include_tasks.isChecked():
            self.add_sheet_to_workbook(wb, "Tasks", '''
                                                    SELECT t.id,
                                                           p.name     as project,
                                                           t.title,
                                                           t.status,
                                                           t.priority,
                                                           t.due_date,
                                                           u.username as assigned_to
                                                    FROM tasks t
                                                             LEFT JOIN projects p ON t.project_id = p.id
                                                             LEFT JOIN users u ON t.assigned_to = u.id
                                                    WHERE t.created_at BETWEEN ? AND ?
                                                    ORDER BY t.due_date, t.priority DESC
                                                    ''', (from_date, to_date),
                                       ["ID", "Project", "Title", "Status", "Priority", "Due Date", "Assigned To"])

        if self.include_bugs.isChecked():
            self.add_sheet_to_workbook(wb, "Bugs", '''
                                                   SELECT b.id,
                                                          p.name as project,
                                                          b.title,
                                                          b.severity,
                                                          b.status,
                                                          b.created_at
                                                   FROM bugs b
                                                            LEFT JOIN projects p ON b.project_id = p.id
                                                   WHERE b.created_at BETWEEN ? AND ?
                                                   ORDER BY b.severity DESC, b.created_at DESC
                                                   ''', (from_date, to_date),
                                       ["ID", "Project", "Title", "Severity", "Status", "Reported On"])

        if self.include_activities.isChecked():
            self.add_sheet_to_workbook(wb, "Activities", '''
                                                         SELECT id,
                                                                title,
                                                                type,
                                                                status,
                                                                created_at
                                                         FROM extra_activities
                                                         WHERE created_at BETWEEN ? AND ?
                                                         ORDER BY created_at DESC
                                                         ''', (from_date, to_date),
                                       ["ID", "Title", "Type", "Status", "Requested On"])

        # Save the workbook
        report_name = f"Comprehensive_Report_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        report_path = os.path.join(os.getcwd(), "reports", report_name)

        # Ensure reports directory exists
        os.makedirs(os.path.dirname(report_path), exist_ok=True)
        wb.save(report_path)

        # Show confirmation
        QMessageBox.information(
            self,
            "Report Generated",
            f"Comprehensive report saved to:\n{report_path}"
        )

    def add_sheet_to_workbook(self, workbook, sheet_name, query, params, headers):
        """Add a sheet with data to the workbook"""
        cursor = self.db['sqlite'].cursor()
        cursor.execute(query, params)
        data = cursor.fetchall()

        ws = workbook.create_sheet(title=sheet_name)

        # Add headers
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}1"] = header
            ws[f"{col_letter}1"].font = Font(bold=True)

        # Add data
        for row_num, row_data in enumerate(data, 2):
            for col_num, cell_value in enumerate(row_data, 1):
                col_letter = get_column_letter(col_num)
                ws[f"{col_letter}{row_num}"] = cell_value

        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

    def generate_excel_report(self, report_name, headers, data):
        """Generate a simple Excel report with one sheet"""
        wb = Workbook()
        ws = wb.active
        ws.title = report_name[:31]  # Excel sheet name limit

        # Add headers
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}1"] = header
            ws[f"{col_letter}1"].font = Font(bold=True)

        # Add data
        for row_num, row_data in enumerate(data, 2):
            for col_num, cell_value in enumerate(row_data, 1):
                col_letter = get_column_letter(col_num)
                ws[f"{col_letter}{row_num}"] = cell_value

        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

        # Save the workbook
        report_file = f"{report_name.replace(' ', '_')}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        report_path = os.path.join(os.getcwd(), "reports", report_file)

        # Ensure reports directory exists
        os.makedirs(os.path.dirname(report_path), exist_ok=True)
        wb.save(report_path)